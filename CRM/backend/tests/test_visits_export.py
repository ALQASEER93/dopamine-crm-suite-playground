from __future__ import annotations

from datetime import date
import io

from fastapi.testclient import TestClient


def test_visits_export_csv(client: TestClient, auth_headers: dict[str, str]) -> None:
    resp = client.get("/api/v1/visits/export", headers=auth_headers)
    assert resp.status_code == 200, resp.text
    assert "text/csv" in resp.headers.get("content-type", "")
    assert "visits.csv" in resp.headers.get("content-disposition", "")


def test_visits_export_excel(client: TestClient, auth_headers: dict[str, str]) -> None:
    resp = client.get("/api/v1/visits/export/excel", headers=auth_headers)
    assert resp.status_code == 200, resp.text
    assert (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        in resp.headers.get("content-type", "")
    )
    disposition = resp.headers.get("content-disposition", "")
    assert "visits_" in disposition
    assert ".xlsx" in disposition

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(resp.content))
    assert "Visits" in wb.sheetnames


def test_visits_export_excel_sanitizes_formula_like_text(
    client: TestClient, auth_headers: dict[str, str]
) -> None:
    doctor_resp = client.post(
        "/api/v1/doctors",
        headers=auth_headers,
        json={"name": "Dr. Export Formula", "specialty": "GP", "area": "Central"},
    )
    assert doctor_resp.status_code in (200, 201), doctor_resp.text
    doctor_id = doctor_resp.json()["id"]

    reps_resp = client.get("/api/v1/reps", headers=auth_headers)
    rep_id = reps_resp.json()[0]["id"]

    payloads = [
        "=HYPERLINK(\"http://malicious\")",
        "   =1+1",
        "'   +SUM(1,2)",
        "\"   -2+3",
        "  '@cmd",
    ]
    visit_ids: list[int] = []
    for note in payloads:
        create_visit_resp = client.post(
            "/api/v1/visits",
            headers=auth_headers,
            json={
                "visit_date": date.today().isoformat(),
                "rep_id": rep_id,
                "doctor_id": doctor_id,
                "notes": note,
            },
        )
        assert create_visit_resp.status_code in (200, 201), create_visit_resp.text
        visit_ids.append(create_visit_resp.json()["id"])

    export_resp = client.get(f"/api/v1/visits/export/excel?rep_id={rep_id}", headers=auth_headers)
    assert export_resp.status_code == 200, export_resp.text

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(export_resp.content), data_only=False)
    ws = wb["Visits"]

    notes_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "ملاحظات":
            notes_col_idx = idx
            break
    assert notes_col_idx is not None

    notes_by_visit: dict[int, object] = {}
    for row in ws.iter_rows(min_row=2):
        visit_id = row[0].value
        if visit_id in visit_ids:
            notes_by_visit[visit_id] = row[notes_col_idx - 1].value

    assert len(notes_by_visit) == len(visit_ids)
    for idx, note in enumerate(payloads):
        stored_note = notes_by_visit[visit_ids[idx]]
        assert stored_note == f"'{note}"
