from __future__ import annotations

from datetime import date, datetime, timedelta, timezone
import logging
import csv
import io

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from fastapi.responses import StreamingResponse
from sqlalchemy import case, func
from sqlalchemy.orm import Session, joinedload, selectinload

from api.v1.utils import DEFAULT_PAGE, DEFAULT_PAGE_SIZE, clamp_page_size, paginate
from api.v1.utils_gps import GPSValidationError, validate_accuracy, validate_geofence, validate_max_distance
from core.config import settings
from core.db import get_db
from core.security import get_current_user, has_any_role, require_roles
from models.crm import Doctor, Pharmacy, User, Visit
from schemas.common import PaginatedResponse
from schemas.crm import VisitCreate, VisitEnd, VisitOut, VisitStart, VisitUpdate
from services.customer_scope import is_customer_assigned_to_rep

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/visits",
    tags=["visits"],
    dependencies=[Depends(get_current_user)],
)

VISIT_STATUSES = {"scheduled", "in_progress", "completed", "cancelled"}
FORMULA_PREFIXES = ("=", "+", "-", "@")
MAX_CLIENT_CLOCK_SKEW = timedelta(minutes=5)


def _validate_lifecycle_timestamp(value: datetime | None, field_name: str) -> datetime:
    timestamp = value or datetime.now(timezone.utc)
    comparable = timestamp if timestamp.tzinfo else timestamp.replace(tzinfo=timezone.utc)
    if comparable.astimezone(timezone.utc) > datetime.now(timezone.utc) + MAX_CLIENT_CLOCK_SKEW:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{field_name} cannot be in the future.",
        )
    return timestamp


def _calculate_duration_seconds(started_at: datetime | None, ended_at: datetime | None) -> int | None:
    if not started_at or not ended_at:
        return None
    if started_at.tzinfo is None and ended_at.tzinfo:
        started_at = started_at.replace(tzinfo=ended_at.tzinfo)
    if ended_at.tzinfo is None and started_at.tzinfo:
        ended_at = ended_at.replace(tzinfo=started_at.tzinfo)
    delta = (ended_at - started_at).total_seconds()
    return max(int(delta), 0)


def _normalize_status_filters(status_filters: list[str] | None) -> list[str]:
    if not status_filters:
        return []
    normalized = []
    for candidate in status_filters:
        if not candidate:
            continue
        value = candidate.strip().lower()
        if value in VISIT_STATUSES:
            normalized.append(value)
    # Preserve order while de-duplicating
    return list(dict.fromkeys(normalized))


def _sync_duration(visit: Visit) -> None:
    duration = _calculate_duration_seconds(visit.started_at, visit.ended_at)
    if duration is not None:
        visit.duration_seconds = duration


def _sanitize_excel_text(value: object | None) -> object | None:
    if not isinstance(value, str) or not value:
        return value
    candidate = value.lstrip()
    while candidate.startswith(("'", '"')):
        candidate = candidate[1:].lstrip()
    if candidate.startswith(FORMULA_PREFIXES):
        return f"'{value}"
    return value


@router.get("", response_model=PaginatedResponse[VisitOut], include_in_schema=False)
@router.get("/", response_model=PaginatedResponse[VisitOut])
def list_visits(
    page: int = Query(DEFAULT_PAGE, ge=1),
    page_size: int = Query(DEFAULT_PAGE_SIZE, ge=1, le=500),
    rep_ids: list[int] | None = Query(default=None, alias="rep_id"),
    doctor_id: int | None = None,
    pharmacy_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    status_filter: list[str] | None = Query(default=None, alias="status"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> PaginatedResponse[VisitOut]:
    query = (
        db.query(Visit)
        .options(joinedload(Visit.doctor), joinedload(Visit.pharmacy), joinedload(Visit.rep))
        .filter(Visit.is_deleted.is_(False))
    )

    effective_rep_ids = rep_ids or []
    if has_any_role(current_user, ["medical_rep"]):
        effective_rep_ids = [current_user.id]

    if effective_rep_ids:
        query = query.filter(Visit.rep_id.in_(effective_rep_ids))
    if doctor_id:
        query = query.filter(Visit.doctor_id == doctor_id)
    if pharmacy_id:
        query = query.filter(Visit.pharmacy_id == pharmacy_id)
    if date_from:
        query = query.filter(Visit.visit_date >= date_from)
    if date_to:
        query = query.filter(Visit.visit_date <= date_to)
    allowed_statuses = _normalize_status_filters(status_filter)
    if allowed_statuses:
        query = query.filter(Visit.status.in_(allowed_statuses))

    page_size = clamp_page_size(page_size)
    visits, total = paginate(
        query.order_by(
            Visit.started_at.desc().nullslast(),
            Visit.visit_date.desc(),
            Visit.id.desc(),
        ),
        page,
        page_size,
    )
    total_pages = max(1, (total + page_size - 1) // page_size)
    return PaginatedResponse(
        data=visits,
        pagination={"page": page, "page_size": page_size, "total": total, "total_pages": total_pages},
    )


@router.get("/export", dependencies=[Depends(require_roles("sales_manager", "admin"))])
def export_visits(
    rep_ids: list[int] | None = Query(default=None, alias="rep_id"),
    doctor_id: int | None = None,
    pharmacy_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    status_filter: list[str] | None = Query(default=None, alias="status"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Response:
    query = (
        db.query(Visit)
        .options(joinedload(Visit.doctor), joinedload(Visit.pharmacy), joinedload(Visit.rep))
        .filter(Visit.is_deleted.is_(False))
    )

    effective_rep_ids = rep_ids or []
    if has_any_role(current_user, ["medical_rep"]):
        effective_rep_ids = [current_user.id]

    if effective_rep_ids:
        query = query.filter(Visit.rep_id.in_(effective_rep_ids))
    if doctor_id:
        query = query.filter(Visit.doctor_id == doctor_id)
    if pharmacy_id:
        query = query.filter(Visit.pharmacy_id == pharmacy_id)
    if date_from:
        query = query.filter(Visit.visit_date >= date_from)
    if date_to:
        query = query.filter(Visit.visit_date <= date_to)
    allowed_statuses = _normalize_status_filters(status_filter)
    if allowed_statuses:
        query = query.filter(Visit.status.in_(allowed_statuses))

    visits = query.order_by(
        Visit.started_at.desc().nullslast(),
        Visit.visit_date.desc(),
        Visit.id.desc(),
    ).all()

    buffer = io.StringIO()
    writer = csv.DictWriter(
        buffer,
        fieldnames=[
            "id",
            "visitDate",
            "status",
            "durationMinutes",
            "repName",
            "repEmail",
            "accountName",
            "accountType",
            "notes",
        ],
    )
    writer.writeheader()

    for visit in visits:
        duration_minutes = None
        if visit.duration_seconds is not None:
            duration_minutes = round(visit.duration_seconds / 60, 2)
        elif visit.started_at and visit.ended_at:
            duration_minutes = round((visit.ended_at - visit.started_at).total_seconds() / 60, 2)

        account_name = None
        account_type = None
        if visit.doctor:
            account_name = visit.doctor.name
            account_type = "doctor"
        elif visit.pharmacy:
            account_name = visit.pharmacy.name
            account_type = "pharmacy"

        writer.writerow(
            {
                "id": visit.id,
                "visitDate": visit.visit_date.isoformat() if visit.visit_date else "",
                "status": visit.status,
                "durationMinutes": duration_minutes if duration_minutes is not None else "",
                "repName": visit.rep.name if visit.rep else "",
                "repEmail": visit.rep.email if visit.rep else "",
                "accountName": account_name or "",
                "accountType": account_type or "",
                "notes": visit.notes or "",
            }
        )

    headers = {"Content-Disposition": 'attachment; filename="visits.csv"'}
    return Response(content=buffer.getvalue(), media_type="text/csv", headers=headers)


@router.get("/export/excel", dependencies=[Depends(require_roles("sales_manager", "admin"))])
def export_visits_excel(
    rep_ids: list[int] | None = Query(default=None, alias="rep_id"),
    doctor_id: int | None = None,
    pharmacy_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    status_filter: list[str] | None = Query(default=None, alias="status"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """Export visits to Excel (Arabic RTL headers)."""
    try:
        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel export requires openpyxl. Install with: pip install openpyxl"
        )

    query = (
        db.query(Visit)
        .options(selectinload(Visit.doctor), selectinload(Visit.pharmacy), selectinload(Visit.rep))
        .filter(Visit.is_deleted.is_(False))
    )

    effective_rep_ids = rep_ids or []
    if has_any_role(current_user, ["medical_rep"]):
        effective_rep_ids = [current_user.id]

    if effective_rep_ids:
        query = query.filter(Visit.rep_id.in_(effective_rep_ids))
    if doctor_id:
        query = query.filter(Visit.doctor_id == doctor_id)
    if pharmacy_id:
        query = query.filter(Visit.pharmacy_id == pharmacy_id)
    if date_from:
        query = query.filter(Visit.visit_date >= date_from)
    if date_to:
        query = query.filter(Visit.visit_date <= date_to)
    allowed_statuses = _normalize_status_filters(status_filter)
    if allowed_statuses:
        query = query.filter(Visit.status.in_(allowed_statuses))

    query = query.order_by(
        Visit.started_at.desc().nullslast(),
        Visit.visit_date.desc(),
        Visit.id.desc(),
    )

    def _iso(dt: datetime | None) -> str:
        return dt.isoformat() if dt else ""

    def _safe_str(value: object | None) -> str:
        if value is None:
            return ""
        return str(value)

    columns = [
        ("رقم الزيارة", lambda v: v.id),
        ("تاريخ الزيارة", lambda v: v.visit_date.isoformat() if v.visit_date else ""),
        ("الحالة", lambda v: v.status),
        ("رقم المندوب", lambda v: v.rep_id),
        ("اسم المندوب", lambda v: v.rep.name if v.rep else ""),
        ("بريد المندوب", lambda v: v.rep.email if v.rep else ""),
        ("رقم الطبيب", lambda v: v.doctor_id or ""),
        ("اسم الطبيب", lambda v: v.doctor.name if v.doctor else ""),
        ("رقم الصيدلية", lambda v: v.pharmacy_id or ""),
        ("اسم الصيدلية", lambda v: v.pharmacy.name if v.pharmacy else ""),
        ("ملاحظات", lambda v: v.notes or ""),
        ("عينات", lambda v: v.samples_given or ""),
        ("إجراء لاحق", lambda v: v.next_action or ""),
        ("تاريخ الإجراء التالي", lambda v: v.next_action_date.isoformat() if v.next_action_date else ""),
        ("بدء الزيارة", lambda v: _iso(v.started_at)),
        ("نهاية الزيارة", lambda v: _iso(v.ended_at)),
        ("المدة (ثانية)", lambda v: v.duration_seconds or ""),
        ("المدة (دقيقة)", lambda v: round(v.duration_seconds / 60, 2) if v.duration_seconds is not None else ""),
        ("خط العرض (بداية)", lambda v: v.start_lat if v.start_lat is not None else ""),
        ("خط الطول (بداية)", lambda v: v.start_lng if v.start_lng is not None else ""),
        ("دقة GPS (بداية-م)", lambda v: v.start_accuracy if v.start_accuracy is not None else ""),
        ("خط العرض (نهاية)", lambda v: v.end_lat if v.end_lat is not None else ""),
        ("خط الطول (نهاية)", lambda v: v.end_lng if v.end_lng is not None else ""),
        ("دقة GPS (نهاية-م)", lambda v: v.end_accuracy if v.end_accuracy is not None else ""),
        ("تاريخ الإنشاء", lambda v: _iso(v.created_at)),
        ("آخر تحديث", lambda v: _iso(v.updated_at)),
    ]

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Visits")
    try:
        ws.sheet_view.rightToLeft = True
    except Exception:  # noqa: BLE001
        pass

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    max_widths = [len(header) for header, _ in columns]

    header_row: list[WriteOnlyCell] = []
    for header, _getter in columns:
        cell = WriteOnlyCell(ws, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        header_row.append(cell)
    ws.append(header_row)

    for visit in query.yield_per(1000):
        row: list[object] = []
        for idx, (_header, getter) in enumerate(columns):
            value = _sanitize_excel_text(getter(visit))
            row.append(value)
            width = len(_safe_str(value))
            if width > max_widths[idx]:
                max_widths[idx] = min(width, 80)
        ws.append(row)

    for idx, width in enumerate(max_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 12), 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"visits_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@router.post(
    "",
    status_code=status.HTTP_201_CREATED,
    response_model=VisitOut,
    dependencies=[Depends(require_roles("sales_manager", "medical_rep", "admin"))],
    include_in_schema=False,
)
@router.post(
    "/",
    status_code=status.HTTP_201_CREATED,
    response_model=VisitOut,
    dependencies=[Depends(require_roles("sales_manager", "medical_rep", "admin"))],
)
def create_visit(
    payload: VisitCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Visit:
    if payload.status != "scheduled":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Visit lifecycle fields can only be changed through start/end endpoints.",
        )
    if payload.doctor_id and not db.get(Doctor, payload.doctor_id):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Doctor not found.")
    if payload.pharmacy_id and not db.get(Pharmacy, payload.pharmacy_id):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Pharmacy not found.")
    rep = db.get(User, payload.rep_id)
    if not rep:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Rep not found.")
    if has_any_role(current_user, ["medical_rep"]) and payload.rep_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed to assign another rep.")
    customer_type = "doctor" if payload.doctor_id else "pharmacy"
    customer_id = payload.doctor_id or payload.pharmacy_id
    if has_any_role(current_user, ["medical_rep"]) and not is_customer_assigned_to_rep(
        db,
        rep_id=current_user.id,
        customer_type=customer_type,
        customer_id=customer_id,
    ):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Customer is not assigned to the current representative.",
        )

    visit = Visit(**payload.model_dump())
    db.add(visit)
    db.commit()
    db.refresh(visit)
    return visit


@router.get("/summary")
def visits_summary(
    rep_ids: list[int] | None = Query(default=None, alias="rep_id"),
    doctor_id: int | None = None,
    pharmacy_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    status_filter: list[str] | None = Query(default=None, alias="status"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    base_query = db.query(Visit).filter(Visit.is_deleted.is_(False))
    effective_rep_ids = rep_ids or []
    if has_any_role(current_user, ["medical_rep"]):
        effective_rep_ids = [current_user.id]

    if effective_rep_ids:
        base_query = base_query.filter(Visit.rep_id.in_(effective_rep_ids))
    if doctor_id:
        base_query = base_query.filter(Visit.doctor_id == doctor_id)
    if pharmacy_id:
        base_query = base_query.filter(Visit.pharmacy_id == pharmacy_id)
    if date_from:
        base_query = base_query.filter(Visit.visit_date >= date_from)
    if date_to:
        base_query = base_query.filter(Visit.visit_date <= date_to)
    allowed_statuses = _normalize_status_filters(status_filter)
    if allowed_statuses:
        base_query = base_query.filter(Visit.status.in_(allowed_statuses))

    total_visits = base_query.count()
    completed_visits = base_query.filter(Visit.status == "completed").count()
    scheduled_visits = base_query.filter(Visit.status == "scheduled").count()
    cancelled_visits = base_query.filter(Visit.status == "cancelled").count()
    in_progress_visits = base_query.filter(Visit.status == "in_progress").count()

    duration_sum, duration_count = (
        base_query.filter(Visit.duration_seconds.isnot(None))
        .with_entities(func.sum(Visit.duration_seconds), func.count(Visit.id))
        .first()
    )
    avg_duration_minutes = (
        round(float(duration_sum or 0) / 60 / max(duration_count, 1), 2) if duration_count else 0.0
    )

    last_visit = (
        base_query.order_by(
            Visit.ended_at.desc().nullslast(),
            Visit.started_at.desc().nullslast(),
            Visit.visit_date.desc(),
            Visit.id.desc(),
        )
        .first()
    )
    last_activity = None
    if last_visit:
        last_activity = last_visit.ended_at or last_visit.started_at
        if not last_activity and last_visit.visit_date:
            last_activity = datetime.combine(
                last_visit.visit_date,
                datetime.min.time(),
                tzinfo=timezone.utc,
            )

    rep_rows = (
        base_query.join(User, User.id == Visit.rep_id)
        .with_entities(
            Visit.rep_id,
            User.name,
            User.email,
            func.count(Visit.id),
            func.sum(case((Visit.status == "completed", 1), else_=0)),
            func.sum(Visit.duration_seconds),
            func.max(func.coalesce(Visit.ended_at, Visit.started_at, Visit.visit_date)),
        )
        .group_by(Visit.rep_id, User.name, User.email)
        .all()
    )

    visits_by_rep = []
    for rep_id, rep_name, rep_email, total, completed_count, duration_total, last_seen in rep_rows:
        avg_minutes = None
        if duration_total:
            avg_minutes = round(float(duration_total) / 60 / max(total, 1), 2)
        if isinstance(last_seen, date) and not isinstance(last_seen, datetime):
            last_seen_str = datetime.combine(last_seen, datetime.min.time(), tzinfo=timezone.utc).isoformat()
        elif isinstance(last_seen, datetime):
            last_seen_str = last_seen.isoformat()
        else:
            last_seen_str = None
        visits_by_rep.append(
            {
                "repId": rep_id,
                "repName": rep_name or rep_email or f"Rep {rep_id}",
                "totalVisits": total,
                "completedVisits": int(completed_count or 0),
                "avgDurationMinutes": avg_minutes,
                "lastVisitAt": last_seen_str,
            }
        )

    completion_rate = round((completed_visits / total_visits) * 100, 2) if total_visits else 0.0
    data = {
        "totalVisits": total_visits,
        "completedVisits": completed_visits,
        "scheduledVisits": scheduled_visits,
        "cancelledVisits": cancelled_visits,
        "inProgressVisits": in_progress_visits,
        "completionRate": completion_rate,
        "avgDurationMinutes": avg_duration_minutes,
        "lastActivityAt": last_activity.isoformat() if isinstance(last_activity, datetime) else None,
        "visitsByRep": visits_by_rep,
    }
    return {"data": data}


def _serialize_dashboard_visit(visit: Visit) -> dict:
    rep = None
    if visit.rep:
        rep = {"id": visit.rep.id, "name": visit.rep.name or visit.rep.email}

    hcp = None
    account = None
    if visit.doctor:
        hcp = {"id": visit.doctor.id, "name": visit.doctor.name, "type": "doctor"}
        account = {"type": "doctor", "id": visit.doctor.id, "name": visit.doctor.name}
    elif visit.pharmacy:
        hcp = {"id": visit.pharmacy.id, "name": visit.pharmacy.name, "type": "pharmacy"}
        account = {"type": "pharmacy", "id": visit.pharmacy.id, "name": visit.pharmacy.name}

    duration_seconds = visit.duration_seconds or _calculate_duration_seconds(visit.started_at, visit.ended_at) or 0
    duration_minutes = round(duration_seconds / 60, 2) if duration_seconds else 0

    return {
        "id": visit.id,
        "visitDate": visit.visit_date.isoformat() if visit.visit_date else None,
        "startedAt": visit.started_at.isoformat() if visit.started_at else None,
        "endedAt": visit.ended_at.isoformat() if visit.ended_at else None,
        "rep": rep,
        "hcp": hcp,
        "account": account,
        "status": visit.status or "scheduled",
        "durationMinutes": duration_minutes,
        "startLocation": {
            "lat": visit.start_lat,
            "lng": visit.start_lng,
            "accuracy": visit.start_accuracy,
        },
        "endLocation": {
            "lat": visit.end_lat,
            "lng": visit.end_lng,
            "accuracy": visit.end_accuracy,
        },
        "notes": visit.notes,
        "samplesGiven": visit.samples_given,
        "nextAction": visit.next_action,
        "nextActionDate": visit.next_action_date.isoformat() if visit.next_action_date else None,
    }


@router.get("/latest")
def latest_visits(
    page_size: int = Query(5, alias="pageSize", ge=1, le=50),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    page_size = clamp_page_size(page_size)
    query = (
        db.query(Visit)
        .options(joinedload(Visit.rep), joinedload(Visit.doctor), joinedload(Visit.pharmacy))
        .filter(Visit.is_deleted.is_(False))
        .order_by(
            Visit.ended_at.desc().nullslast(),
            Visit.started_at.desc().nullslast(),
            Visit.visit_date.desc(),
            Visit.id.desc(),
        )
    )

    if has_any_role(current_user, ["medical_rep"]):
        query = query.filter(Visit.rep_id == current_user.id)

    visits = query.limit(page_size).all()
    data = [_serialize_dashboard_visit(visit) for visit in visits]
    return {"data": data}


def _get_visit(db: Session, visit_id: int) -> Visit:
    visit = (
        db.query(Visit)
        .options(joinedload(Visit.rep), joinedload(Visit.doctor), joinedload(Visit.pharmacy))
        .filter(Visit.id == visit_id, Visit.is_deleted.is_(False))
        .first()
    )
    if not visit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Visit not found.")
    return visit


@router.get("/{visit_id:int}", response_model=VisitOut)
def get_visit(visit_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Visit:
    visit = _get_visit(db, visit_id)
    if has_any_role(current_user, ["medical_rep"]) and visit.rep_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not permitted.")
    return visit


@router.put(
    "/{visit_id:int}",
    response_model=VisitOut,
    dependencies=[Depends(require_roles("sales_manager", "medical_rep", "admin"))],
)
def update_visit(
    visit_id: int,
    payload: VisitUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Visit:
    visit = _get_visit(db, visit_id)
    if has_any_role(current_user, ["medical_rep"]) and visit.rep_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not permitted.")

    updates = payload.model_dump(exclude_unset=True)
    forbidden = {"status", "started_at", "ended_at", "start_lat", "start_lng", "start_accuracy", "end_lat", "end_lng", "end_accuracy", "duration_seconds"}
    if forbidden.intersection(updates.keys()):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Visit lifecycle fields can only be changed through start/end endpoints.",
        )
    if "notes" in updates:
        is_active_window = (
            visit.status == "in_progress"
            and visit.started_at is not None
            and visit.ended_at is None
        )
        if not is_active_window:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Visit notes can only be edited while the visit is active.",
            )
    if "doctor_id" in updates and updates["doctor_id"]:
        if not db.get(Doctor, updates["doctor_id"]):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Doctor not found.")
    if "pharmacy_id" in updates and updates["pharmacy_id"]:
        if not db.get(Pharmacy, updates["pharmacy_id"]):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Pharmacy not found.")
    if "rep_id" in updates and updates["rep_id"]:
        rep = db.get(User, updates["rep_id"])
        if not rep:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Rep not found.")
        if has_any_role(current_user, ["medical_rep"]) and updates["rep_id"] != current_user.id:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not permitted to reassign rep.")

    for key, value in updates.items():
        setattr(visit, key, value)

    if not visit.doctor_id and not visit.pharmacy_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Either doctor_id or pharmacy_id is required.",
        )
    if visit.doctor_id and visit.pharmacy_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Provide only one of doctor_id or pharmacy_id.",
        )
    if has_any_role(current_user, ["medical_rep"]):
        customer_type = "doctor" if visit.doctor_id else "pharmacy"
        customer_id = visit.doctor_id or visit.pharmacy_id
        if not is_customer_assigned_to_rep(
            db,
            rep_id=current_user.id,
            customer_type=customer_type,
            customer_id=customer_id,
        ):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Customer is not assigned to the current representative.",
            )

    db.commit()
    db.refresh(visit)
    return visit


@router.post(
    "/{visit_id:int}/start",
    response_model=VisitOut,
    dependencies=[Depends(require_roles("sales_manager", "medical_rep", "admin"))],
)
def start_visit(
    visit_id: int,
    payload: VisitStart,
    gps_override: bool = Query(default=False, alias="gpsOverride"),
    gps_override_reason: str | None = Query(default=None, alias="gpsOverrideReason"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Visit:
    visit = _get_visit(db, visit_id)
    if has_any_role(current_user, ["medical_rep"]) and visit.rep_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not permitted.")
    if visit.started_at:
        logger.info("Start visit called for already-started visit (visit_id=%s user_id=%s).", visit.id, current_user.id)
        return visit
    if visit.status != "scheduled":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only a scheduled visit can be started.",
        )

    allow_override = (
        gps_override
        and bool(settings.allow_gps_override)
        and has_any_role(current_user, ["admin", "sales_manager"])
    )
    if allow_override:
        normalized_reason = (gps_override_reason or "").strip() or None
        logger.warning(
            "GPS override used for visit start (visit_id=%s user_id=%s reason_supplied=%s).",
            visit.id,
            current_user.id,
            bool(normalized_reason),
        )
    else:
        try:
            validate_accuracy(payload.accuracy)
            # Geofencing validation: ensure rep is near the doctor/pharmacy (feature-flagged in settings).
            if visit.doctor_id and visit.doctor:
                validate_geofence(payload.lat, payload.lng, visit.doctor.latitude, visit.doctor.longitude)
            elif visit.pharmacy_id and visit.pharmacy:
                validate_geofence(payload.lat, payload.lng, visit.pharmacy.latitude, visit.pharmacy.longitude)
        except GPSValidationError as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc

    started_at = _validate_lifecycle_timestamp(payload.started_at, "Visit start time")
    visit.started_at = started_at
    visit.start_lat = payload.lat
    visit.start_lng = payload.lng
    visit.start_accuracy = payload.accuracy
    visit.status = "in_progress"
    _sync_duration(visit)

    db.commit()
    db.refresh(visit)
    return visit


@router.post(
    "/{visit_id:int}/end",
    response_model=VisitOut,
    dependencies=[Depends(require_roles("sales_manager", "medical_rep", "admin"))],
)
def end_visit(
    visit_id: int,
    payload: VisitEnd,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Visit:
    visit = _get_visit(db, visit_id)
    if has_any_role(current_user, ["medical_rep"]) and visit.rep_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not permitted.")
    if visit.ended_at:
        logger.info("End visit called for already-ended visit (visit_id=%s user_id=%s).", visit.id, current_user.id)
        return visit
    if visit.status != "in_progress" or not visit.started_at:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only an active visit can be completed.",
        )

    try:
        validate_accuracy(payload.accuracy)
        validate_max_distance(
            visit.start_lat,
            visit.start_lng,
            payload.lat,
            payload.lng,
        )
    except GPSValidationError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc

    ended_at = _validate_lifecycle_timestamp(payload.ended_at, "Visit end time")
    comparable_started_at = visit.started_at
    comparable_ended_at = ended_at
    if comparable_started_at.tzinfo is None and comparable_ended_at.tzinfo:
        comparable_started_at = comparable_started_at.replace(tzinfo=comparable_ended_at.tzinfo)
    if comparable_ended_at.tzinfo is None and comparable_started_at.tzinfo:
        comparable_ended_at = comparable_ended_at.replace(tzinfo=comparable_started_at.tzinfo)
    if comparable_ended_at <= comparable_started_at:
        if payload.ended_at is not None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Visit end time must be after visit start time.",
            )
        ended_at = visit.started_at + timedelta(microseconds=1)
    visit.ended_at = ended_at
    visit.end_lat = payload.lat
    visit.end_lng = payload.lng
    visit.end_accuracy = payload.accuracy
    visit.status = "completed"
    _sync_duration(visit)

    db.commit()
    db.refresh(visit)
    return visit


@router.delete(
    "/{visit_id:int}",
    status_code=status.HTTP_204_NO_CONTENT,
    response_class=Response,
    dependencies=[Depends(require_roles("sales_manager", "medical_rep", "admin"))],
)
def delete_visit(
    visit_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Response:
    visit = _get_visit(db, visit_id)
    if has_any_role(current_user, ["medical_rep"]) and visit.rep_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not permitted.")
    visit.is_deleted = True
    db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)
