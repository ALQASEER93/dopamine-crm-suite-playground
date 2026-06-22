from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
import csv
import hashlib
import json
from typing import Any, Iterable

from openpyxl import load_workbook
from sqlalchemy import and_
from sqlalchemy.orm import Session

from models.crm import CustomerImportRun, CustomerImportStagingItem, Doctor, Pharmacy


CUSTOMERS_SHEET_NAME = "CRM_Import_Customers"
ROUTE_ASSIGNMENT_BLOCKER = "missing_admin_reviewed_route_assignment"

HEADER_ALIASES = {
    "name": ("name_clean", "name", "Name", "name_original"),
    "customer_type": ("customer_type", "Client Tag", "source_client_tag"),
    "specialty": ("specialty", "Speciality", "source_speciality"),
    "classification": ("classification",),
    "area": ("area_tag", "Area Tag", "area"),
    "city": ("city", "City"),
    "phone": ("phone", "Phone"),
    "email": ("email", "Email"),
    "address": ("formatted_address", "Formatted Address", "clinic"),
    "latitude": ("latitude", "lat"),
    "longitude": ("longitude", "lng", "lon"),
    "location_status": ("location_status",),
    "location_confidence": ("location_confidence",),
    "requires_review": ("requires_review",),
    "review_reason": ("review_reason",),
    "duplicate_group_count": ("duplicate_group_count",),
    "monthly_frequency_target": ("monthly_frequency_target",),
    "assigned_rep": ("assigned_rep", "assigned_rep_email", "rep_email", "rep"),
}

CSV_FIELDS = [
    "customer_type",
    "id",
    "name",
    "specialty",
    "classification",
    "area",
    "city",
    "phone",
    "email",
    "address",
    "latitude",
    "longitude",
]


@dataclass(frozen=True)
class SkippedRows:
    missing_name: int = 0
    missing_customer_type: int = 0
    unsupported_customer_type: int = 0

    def as_api(self) -> dict[str, int]:
        return {
            "missingName": self.missing_name,
            "missingCustomerType": self.missing_customer_type,
            "unsupportedCustomerType": self.unsupported_customer_type,
        }


@dataclass(frozen=True)
class CustomerImportRow:
    row_number: int
    customer_type: str
    name: str
    specialty: str | None
    classification: str | None
    area: str | None
    city: str | None
    phone: str | None
    email: str | None
    address: str | None
    latitude: float | None
    longitude: float | None
    requires_review: bool
    review_reason: str | None
    location_status: str | None
    duplicate_group_count: int
    monthly_frequency_target: int | None
    assigned_rep: str | None


@dataclass(frozen=True)
class WorkbookParseResult:
    rows: list[CustomerImportRow]
    total_rows: int
    skipped: SkippedRows
    unsupported_types: dict[str, int]


@dataclass(frozen=True)
class CustomerPlanItem:
    row: CustomerImportRow
    action: str
    values: dict[str, Any]
    existing_id: int | None


@dataclass(frozen=True)
class CustomerImportPlan:
    items: list[CustomerPlanItem]
    summary: dict[str, Any]
    parse_result: WorkbookParseResult


def _clean_text(value: Any, *, max_length: int | None = None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if max_length and len(text) > max_length:
        return text[:max_length]
    return text


def _clean_classification(value: Any) -> str | None:
    text = _clean_text(value, max_length=1)
    if not text:
        return None
    normalized = text.upper()
    return normalized if normalized in {"A", "B", "C"} else None


def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "y"}


def _as_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(Decimal(str(value).strip()))
    except (InvalidOperation, ValueError):
        return None


def _as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _normalize_customer_type(value: Any) -> str | None:
    raw = str(value or "").strip().lower()
    if raw in {"doctor", "doctor/hcp", "hcp", "physician", "dr"}:
        return "doctor"
    if raw in {"pharmacy", "pharmacy/hco", "hco", "pharmacies"}:
        return "pharmacy"
    return None


def _trusted_coordinates(row: dict[str, Any]) -> tuple[float | None, float | None]:
    lat = _as_float(row.get("latitude"))
    lng = _as_float(row.get("longitude"))
    if lat is None or lng is None:
        return None, None
    if not (-90 <= lat <= 90 and -180 <= lng <= 180):
        return None, None

    status = str(row.get("location_status") or "").strip().lower()
    confidence = _as_float(row.get("location_confidence"))
    if status in {"geocoded", "verified", "trusted"} and confidence is not None and confidence >= 0.8:
        return lat, lng
    return None, None


def _header_index(headers: Iterable[Any]) -> dict[str, int]:
    return {str(header).strip(): index for index, header in enumerate(headers) if header is not None}


def _value(row: tuple[Any, ...], header_map: dict[str, int], canonical: str) -> Any:
    for alias in HEADER_ALIASES[canonical]:
        index = header_map.get(alias)
        if index is not None and index < len(row):
            return row[index]
    return None


def _workbook_rows(content: bytes) -> WorkbookParseResult:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    if CUSTOMERS_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Workbook must include sheet {CUSTOMERS_SHEET_NAME}.")

    worksheet = workbook[CUSTOMERS_SHEET_NAME]
    iterator = worksheet.iter_rows(values_only=True)
    try:
        headers = next(iterator)
    except StopIteration as exc:
        raise ValueError("Workbook customer sheet is empty.") from exc

    header_map = _header_index(headers)
    missing = [
        canonical
        for canonical in ("name", "customer_type")
        if not any(alias in header_map for alias in HEADER_ALIASES[canonical])
    ]
    if missing:
        raise ValueError(f"Workbook is missing required columns: {', '.join(missing)}.")

    rows: list[CustomerImportRow] = []
    total_rows = 0
    missing_name = 0
    missing_customer_type = 0
    unsupported_customer_type = 0
    unsupported_types: dict[str, int] = {}

    for row_number, raw_row in enumerate(iterator, start=2):
        if not any(cell is not None and str(cell).strip() for cell in raw_row):
            continue
        total_rows += 1
        row = {key: _value(raw_row, header_map, key) for key in HEADER_ALIASES}
        name = _clean_text(row["name"], max_length=150)
        raw_type = _clean_text(row["customer_type"], max_length=100)
        customer_type = _normalize_customer_type(raw_type)
        if not name:
            missing_name += 1
            continue
        if not raw_type:
            missing_customer_type += 1
            continue
        if not customer_type:
            unsupported_customer_type += 1
            unsupported_types[raw_type] = unsupported_types.get(raw_type, 0) + 1
            continue

        lat, lng = _trusted_coordinates(row)
        rows.append(
            CustomerImportRow(
                row_number=row_number,
                customer_type=customer_type,
                name=name,
                specialty=_clean_text(row["specialty"], max_length=150),
                classification=_clean_classification(row["classification"]),
                area=_clean_text(row["area"], max_length=100),
                city=_clean_text(row["city"], max_length=100),
                phone=_clean_text(row["phone"], max_length=50),
                email=_clean_text(row["email"], max_length=255),
                address=_clean_text(row["address"], max_length=255),
                latitude=lat,
                longitude=lng,
                requires_review=_as_bool(row["requires_review"]),
                review_reason=_clean_text(row["review_reason"], max_length=500),
                location_status=_clean_text(row["location_status"], max_length=50),
                duplicate_group_count=_as_int(row["duplicate_group_count"]) or 0,
                monthly_frequency_target=_as_int(row["monthly_frequency_target"]),
                assigned_rep=_clean_text(row["assigned_rep"], max_length=255),
            )
        )

    return WorkbookParseResult(
        rows=rows,
        total_rows=total_rows,
        skipped=SkippedRows(
            missing_name=missing_name,
            missing_customer_type=missing_customer_type,
            unsupported_customer_type=unsupported_customer_type,
        ),
        unsupported_types=dict(sorted(unsupported_types.items())),
    )


def _non_empty_differs(instance: Any, values: dict[str, Any]) -> bool:
    return any(value is not None and getattr(instance, field) != value for field, value in values.items())


def _apply_non_empty(instance: Any, values: dict[str, Any]) -> bool:
    changed = False
    for field, value in values.items():
        if value is None:
            continue
        if getattr(instance, field) != value:
            setattr(instance, field, value)
            changed = True
    return changed


def _doctor_identity_filter(row: CustomerImportRow):
    return and_(Doctor.name == row.name, Doctor.clinic == row.address, Doctor.area == row.area)


def _pharmacy_identity_filter(row: CustomerImportRow):
    return and_(Pharmacy.name == row.name, Pharmacy.city == row.city, Pharmacy.area == row.area)


def _find_doctor(db: Session, row: CustomerImportRow) -> Doctor | None:
    doctor = db.query(Doctor).filter(_doctor_identity_filter(row)).order_by(Doctor.id.asc()).first()
    if doctor or row.duplicate_group_count > 1:
        return doctor
    return db.query(Doctor).filter(Doctor.name == row.name, Doctor.area == row.area).order_by(Doctor.id.asc()).first()


def _find_pharmacy(db: Session, row: CustomerImportRow) -> Pharmacy | None:
    pharmacy = db.query(Pharmacy).filter(_pharmacy_identity_filter(row)).order_by(Pharmacy.id.asc()).first()
    if pharmacy or row.duplicate_group_count > 1:
        return pharmacy
    return db.query(Pharmacy).filter(Pharmacy.name == row.name, Pharmacy.area == row.area).order_by(Pharmacy.id.asc()).first()


def _doctor_values(row: CustomerImportRow) -> dict[str, Any]:
    return {
        "specialty": row.specialty,
        "clinic": row.address,
        "area": row.area,
        "city": row.city,
        "classification": row.classification,
        "phone": row.phone,
        "email": row.email,
        "latitude": row.latitude,
        "longitude": row.longitude,
    }


def _pharmacy_values(row: CustomerImportRow) -> dict[str, Any]:
    return {
        "area": row.area,
        "city": row.city,
        "segment": row.specialty,
        "phone": row.phone,
        "email": row.email,
        "latitude": row.latitude,
        "longitude": row.longitude,
    }


def _row_hash(row: CustomerImportRow) -> str:
    payload = {
        "rowNumber": row.row_number,
        "type": row.customer_type,
        "name": row.name,
        "area": row.area,
        "city": row.city,
        "address": row.address,
    }
    return hashlib.sha256(json.dumps(payload, sort_keys=True, ensure_ascii=False).encode("utf-8")).hexdigest()


def _empty_summary(parse_result: WorkbookParseResult, *, dry_run: bool) -> dict[str, Any]:
    return {
        "dryRun": dry_run,
        "sheet": CUSTOMERS_SHEET_NAME,
        "totalRows": parse_result.total_rows,
        "sourceRows": len(parse_result.rows),
        "totalParsedRows": len(parse_result.rows),
        "doctors": 0,
        "pharmacies": 0,
        "created": 0,
        "updated": 0,
        "unchanged": 0,
        "requiresReview": 0,
        "duplicateReviewNeeded": 0,
        "withTrustedCoordinates": 0,
        "monthlyFrequencyTargets": {},
        "areaCount": 0,
        "skipped": parse_result.skipped.as_api(),
        "unsupportedTypes": parse_result.unsupported_types,
        "routeFrequencyAlignment": {
            "status": "pending_admin_review",
            "routeAccountsCreated": 0,
            "assignmentReadyForReview": 0,
            "assignmentMissingSource": 0,
            "pendingAssignments": 0,
            "message": "لم يتم إنشاء تكليفات مسارات تلقائياً. يحتاج الأدمن إلى مراجعة المندوب والتكرار قبل اعتماد المسارات.",
        },
        "audit": {
            "persisted": False,
            "runId": None,
            "status": "planned" if dry_run else "pending_apply",
        },
    }


def build_customer_import_plan(db: Session, content: bytes, *, dry_run: bool = True) -> CustomerImportPlan:
    parse_result = _workbook_rows(content)
    summary = _empty_summary(parse_result, dry_run=dry_run)
    areas: set[str] = set()
    items: list[CustomerPlanItem] = []

    with db.no_autoflush:
        for row in parse_result.rows:
            summary["doctors" if row.customer_type == "doctor" else "pharmacies"] += 1
            if row.requires_review:
                summary["requiresReview"] += 1
            if row.duplicate_group_count > 1:
                summary["duplicateReviewNeeded"] += 1
            if row.latitude is not None and row.longitude is not None:
                summary["withTrustedCoordinates"] += 1
            if row.area:
                areas.add(row.area)
            if row.monthly_frequency_target is not None:
                key = str(row.monthly_frequency_target)
                summary["monthlyFrequencyTargets"][key] = summary["monthlyFrequencyTargets"].get(key, 0) + 1

            if row.assigned_rep:
                summary["routeFrequencyAlignment"]["assignmentReadyForReview"] += 1
            else:
                summary["routeFrequencyAlignment"]["assignmentMissingSource"] += 1
            summary["routeFrequencyAlignment"]["pendingAssignments"] += 1

            if row.customer_type == "doctor":
                existing = _find_doctor(db, row)
                values = _doctor_values(row)
            else:
                existing = _find_pharmacy(db, row)
                values = _pharmacy_values(row)

            if existing:
                action = "updated" if _non_empty_differs(existing, values) else "unchanged"
                existing_id = existing.id
            else:
                action = "created"
                existing_id = None
            summary[action] += 1
            items.append(CustomerPlanItem(row=row, action=action, values=values, existing_id=existing_id))

    summary["areaCount"] = len(areas)
    if summary["routeFrequencyAlignment"]["assignmentMissingSource"]:
        summary["routeFrequencyAlignment"]["status"] = "blocked_missing_assignment_source"
    return CustomerImportPlan(items=items, summary=summary, parse_result=parse_result)


def _audit_from_summary(
    *,
    actor_user_id: int,
    original_filename: str,
    content_hash: str,
    file_size: int,
    summary: dict[str, Any],
    status: str,
    error_summary: str | None = None,
) -> CustomerImportRun:
    skipped = summary["skipped"]
    return CustomerImportRun(
        actor_user_id=actor_user_id,
        dry_run=bool(summary["dryRun"]),
        original_filename=original_filename,
        content_hash=content_hash,
        file_size=file_size,
        source_sheet=summary["sheet"],
        total_parsed_rows=summary["totalParsedRows"],
        doctors_count=summary["doctors"],
        pharmacies_count=summary["pharmacies"],
        created_count=summary["created"],
        updated_count=summary["updated"],
        unchanged_count=summary["unchanged"],
        skipped_missing_name_count=skipped["missingName"],
        skipped_missing_type_count=skipped["missingCustomerType"],
        skipped_unsupported_type_count=skipped["unsupportedCustomerType"],
        review_needed_count=summary["requiresReview"],
        duplicate_review_count=summary["duplicateReviewNeeded"],
        with_trusted_coordinates_count=summary["withTrustedCoordinates"],
        route_assignment_pending_count=summary["routeFrequencyAlignment"]["pendingAssignments"],
        status=status,
        skipped_counts_json=json.dumps(
            {
                "skipped": skipped,
                "unsupportedTypes": summary["unsupportedTypes"],
            },
            ensure_ascii=False,
            sort_keys=True,
        ),
        error_summary=error_summary,
    )


def _record_failed_audit(
    db: Session,
    *,
    actor_user_id: int,
    original_filename: str,
    content_hash: str,
    file_size: int,
    summary: dict[str, Any],
    exc: Exception,
) -> None:
    db.rollback()
    try:
        failed = _audit_from_summary(
            actor_user_id=actor_user_id,
            original_filename=original_filename,
            content_hash=content_hash,
            file_size=file_size,
            summary=summary,
            status="failed",
            error_summary=str(exc)[:500],
        )
        db.add(failed)
        db.commit()
    except Exception:  # noqa: BLE001
        db.rollback()


def _add_staging_item(
    db: Session,
    *,
    import_run_id: int,
    item: CustomerPlanItem,
    doctor_id: int | None,
    pharmacy_id: int | None,
) -> None:
    assignment_status = "pending_admin_review" if item.row.assigned_rep else "blocked_missing_assignment_source"
    db.add(
        CustomerImportStagingItem(
            import_run_id=import_run_id,
            row_number=item.row.row_number,
            row_hash=_row_hash(item.row),
            customer_type=item.row.customer_type,
            doctor_id=doctor_id,
            pharmacy_id=pharmacy_id,
            import_action=item.action,
            monthly_frequency_target=item.row.monthly_frequency_target,
            requires_review=item.row.requires_review or item.row.duplicate_group_count > 1,
            review_reason=item.row.review_reason,
            location_status=item.row.location_status,
            assignment_status=assignment_status,
            assignment_blocker=None if item.row.assigned_rep else ROUTE_ASSIGNMENT_BLOCKER,
        )
    )


def _apply_customer_import_plan(
    db: Session,
    plan: CustomerImportPlan,
    *,
    actor_user_id: int,
    original_filename: str,
    content_hash: str,
    file_size: int,
) -> dict[str, Any]:
    summary = dict(plan.summary)
    summary["dryRun"] = False
    summary["audit"] = {"persisted": False, "runId": None, "status": "pending_apply"}

    try:
        audit = _audit_from_summary(
            actor_user_id=actor_user_id,
            original_filename=original_filename,
            content_hash=content_hash,
            file_size=file_size,
            summary=summary,
            status="applied",
        )
        db.add(audit)
        db.flush()

        for item in plan.items:
            doctor_id = None
            pharmacy_id = None
            if item.row.customer_type == "doctor":
                if item.existing_id:
                    doctor = db.get(Doctor, item.existing_id)
                    if doctor is None:
                        raise ValueError("Planned doctor no longer exists.")
                    _apply_non_empty(doctor, item.values)
                else:
                    doctor = Doctor(name=item.row.name, **item.values)
                    db.add(doctor)
                    db.flush()
                doctor_id = doctor.id
            else:
                if item.existing_id:
                    pharmacy = db.get(Pharmacy, item.existing_id)
                    if pharmacy is None:
                        raise ValueError("Planned pharmacy no longer exists.")
                    _apply_non_empty(pharmacy, item.values)
                else:
                    pharmacy = Pharmacy(name=item.row.name, **item.values)
                    db.add(pharmacy)
                    db.flush()
                pharmacy_id = pharmacy.id

            _add_staging_item(
                db,
                import_run_id=audit.id,
                item=item,
                doctor_id=doctor_id,
                pharmacy_id=pharmacy_id,
            )

        db.commit()
        summary["audit"] = {"persisted": True, "runId": audit.id, "status": "applied"}
        return summary
    except Exception as exc:  # noqa: BLE001
        _record_failed_audit(
            db,
            actor_user_id=actor_user_id,
            original_filename=original_filename,
            content_hash=content_hash,
            file_size=file_size,
            summary=summary,
            exc=exc,
        )
        raise


def import_customers_from_workbook(
    db: Session,
    content: bytes,
    *,
    dry_run: bool,
    actor_user_id: int,
    original_filename: str,
) -> dict[str, Any]:
    content_hash = hashlib.sha256(content).hexdigest()
    plan = build_customer_import_plan(db, content, dry_run=dry_run)
    if dry_run:
        summary = dict(plan.summary)
        summary["contentHash"] = content_hash
        return summary

    summary = _apply_customer_import_plan(
        db,
        plan,
        actor_user_id=actor_user_id,
        original_filename=original_filename,
        content_hash=content_hash,
        file_size=len(content),
    )
    summary["contentHash"] = content_hash
    return summary


def export_customers_csv(db: Session) -> str:
    buffer = StringIO()
    writer = csv.DictWriter(buffer, fieldnames=CSV_FIELDS)
    writer.writeheader()

    for doctor in db.query(Doctor).order_by(Doctor.name.asc(), Doctor.id.asc()).all():
        writer.writerow(
            {
                "customer_type": "doctor",
                "id": doctor.id,
                "name": doctor.name,
                "specialty": doctor.specialty,
                "classification": doctor.classification,
                "area": doctor.area,
                "city": doctor.city,
                "phone": doctor.phone,
                "email": doctor.email,
                "address": doctor.clinic,
                "latitude": doctor.latitude,
                "longitude": doctor.longitude,
            }
        )

    for pharmacy in db.query(Pharmacy).order_by(Pharmacy.name.asc(), Pharmacy.id.asc()).all():
        writer.writerow(
            {
                "customer_type": "pharmacy",
                "id": pharmacy.id,
                "name": pharmacy.name,
                "specialty": pharmacy.segment,
                "classification": "",
                "area": pharmacy.area,
                "city": pharmacy.city,
                "phone": pharmacy.phone,
                "email": pharmacy.email,
                "address": "",
                "latitude": pharmacy.latitude,
                "longitude": pharmacy.longitude,
            }
        )

    return "\ufeff" + buffer.getvalue()
