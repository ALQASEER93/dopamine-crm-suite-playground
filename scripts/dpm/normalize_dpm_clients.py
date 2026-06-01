from __future__ import annotations

import json
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
RUNS_ROOT = REPO_ROOT / "docs" / "_runs"
EXPECTED_COLUMNS = [
    "Name",
    "Representative Name",
    "Area Tag",
    "Client Tag",
    "Speciality",
    "Phone",
    "Latitude",
    "Longitude",
    "Verified",
    "Created At",
    "Updated At",
]
CLIENT_SOURCES = [
    REPO_ROOT / "ALQASEER-PWA" / "hcps.xlsx",
    REPO_ROOT / "CRM" / "hcps.xlsx",
    REPO_ROOT / "CLIENT LIST - DOPAMINE.xlsx",
    REPO_ROOT / "CLIENT LIST - Dopamine.xlsx",
    REPO_ROOT / "CLIENT LIST - DOPAMINE.csv",
    REPO_ROOT / "clients.json",
]


def cell_value(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def infer_type(client_tag: str | None, speciality: str | None) -> str:
    tag = (client_tag or "").strip().lower()
    spec = (speciality or "").strip().lower()
    if tag == "pharmacy" or spec == "pharmacy":
        return "pharmacy"
    if tag in {"a", "b", "c"} or spec:
        return "doctor"
    return "unknown"


def priority_from_tag(client_tag: str | None) -> str | None:
    tag = (client_tag or "").strip().upper()
    return tag if tag in {"A", "B", "C"} else None


def has_coordinate(lat: Any, lng: Any) -> bool:
    try:
        float(lat)
        float(lng)
        return True
    except (TypeError, ValueError):
        return False


def workbook_rows(path: Path) -> tuple[str, list[str], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header_values = next(rows, ())
    headers = [cell_value(value) or f"column_{index + 1}" for index, value in enumerate(header_values)]
    records: list[dict[str, Any]] = []
    for row in rows:
        record = {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}
        if any(cell_value(value) for value in record.values()):
            records.append(record)
    return sheet.title, headers, records


def summarize_source(path: Path) -> dict[str, Any]:
    sheet, columns, rows = workbook_rows(path)
    type_counts: Counter[str] = Counter()
    priority_counts: Counter[str] = Counter()
    rep_counts: Counter[str] = Counter()
    area_counts: Counter[str] = Counter()
    verified_counts: Counter[str] = Counter()
    missing_name = 0
    with_location = 0

    for row in rows:
        name = cell_value(row.get("Name"))
        if not name:
            missing_name += 1
        client_type = infer_type(cell_value(row.get("Client Tag")), cell_value(row.get("Speciality")))
        type_counts[client_type] += 1
        priority_counts[priority_from_tag(cell_value(row.get("Client Tag"))) or "unclassified"] += 1
        rep_counts[cell_value(row.get("Representative Name")) or "unassigned"] += 1
        area_counts[cell_value(row.get("Area Tag")) or "unassigned"] += 1
        verified_counts[cell_value(row.get("Verified")) or "blank"] += 1
        if has_coordinate(row.get("Latitude"), row.get("Longitude")):
            with_location += 1

    return {
        "file": str(path.relative_to(REPO_ROOT)).replace("\\", "/"),
        "sheet": sheet,
        "rowCount": len(rows),
        "columns": columns,
        "expectedColumnsPresent": {column: column in columns for column in EXPECTED_COLUMNS},
        "doctorCount": type_counts["doctor"],
        "pharmacyCount": type_counts["pharmacy"],
        "unknownTypeCount": type_counts["unknown"],
        "priorityCounts": dict(priority_counts),
        "representativeCount": len(rep_counts),
        "topRepresentativesRedacted": [
            {"rank": index + 1, "rowCount": count}
            for index, (_rep, count) in enumerate(rep_counts.most_common(8))
        ],
        "areaCount": len(area_counts),
        "topAreasRedacted": [
            {"rank": index + 1, "rowCount": count}
            for index, (_area, count) in enumerate(area_counts.most_common(8))
        ],
        "verifiedCounts": dict(verified_counts),
        "withLocationCount": with_location,
        "withoutLocationCount": max(0, len(rows) - with_location),
        "missingNameCount": missing_name,
        "normalizedCustomerFields": [
            "customer_type",
            "name",
            "specialty_or_category",
            "territory",
            "area",
            "assigned_rep",
            "phone",
            "latitude",
            "longitude",
            "monthly_frequency_target",
            "priority",
            "verified",
            "notes",
        ],
        "privacy": "No names, phone numbers, addresses, or exact coordinates are written to this summary.",
    }


def write_markdown(summary: dict[str, Any], output_path: Path) -> None:
    lines = [
        "# DPM Client Import Summary",
        "",
        f"- Generated at: {summary['generatedAt']}",
        f"- Status: {summary['status']}",
        f"- Source files found: {len(summary['sources'])}",
        "- Privacy: raw names, phones, addresses, and coordinates are not included in this artifact.",
        "",
    ]
    if summary["sources"]:
        for source in summary["sources"]:
            lines.extend(
                [
                    f"## {source['file']}",
                    f"- Sheet: {source['sheet']}",
                    f"- Rows: {source['rowCount']}",
                    f"- Doctors/HCPs inferred: {source['doctorCount']}",
                    f"- Pharmacies/HCOs inferred: {source['pharmacyCount']}",
                    f"- Unknown type rows: {source['unknownTypeCount']}",
                    f"- Representatives: {source['representativeCount']} (names redacted)",
                    f"- Areas/territories: {source['areaCount']} (names redacted)",
                    f"- Rows with coordinates: {source['withLocationCount']}",
                    f"- Rows without coordinates: {source['withoutLocationCount']}",
                    f"- Missing names: {source['missingNameCount']}",
                    f"- Columns: {', '.join(source['columns'])}",
                    "",
                ]
            )
    else:
        lines.extend(
            [
                "## OWNER_ACTION",
                "Place the real DOPAMINE client workbook in a safe repo-local path such as `ALQASEER-PWA/hcps.xlsx` or upload it to Codex, then rerun the bridge.",
                "Do not invent production doctors/pharmacies.",
                "",
            ]
        )
    output_path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: normalize_dpm_clients.py <docs/_runs/run_timestamp>", file=sys.stderr)
        return 2
    run_dir = Path(sys.argv[1]).resolve()
    if not str(run_dir).startswith(str(RUNS_ROOT.resolve())):
        print("Run directory must be under docs/_runs.", file=sys.stderr)
        return 2

    (run_dir / "json").mkdir(parents=True, exist_ok=True)
    (run_dir / "artifacts").mkdir(parents=True, exist_ok=True)

    sources = []
    for source in CLIENT_SOURCES:
        if source.exists() and source.suffix.lower() == ".xlsx":
            sources.append(summarize_source(source))

    summary = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "status": "FOUND_AND_NORMALIZED" if sources else "CLIENT_FILE_NOT_FOUND",
        "sources": sources,
        "ownerAction": None
        if sources
        else "Place the real DOPAMINE client list in ALQASEER-PWA/hcps.xlsx or CRM/hcps.xlsx and rerun the bridge.",
    }
    (run_dir / "json" / "client_import_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    write_markdown(summary, run_dir / "artifacts" / "CLIENT_IMPORT_SUMMARY.md")
    print(json.dumps({"status": summary["status"], "sources": len(sources)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
